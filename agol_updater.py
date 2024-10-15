# pylint: disable=line-too-long
# pylint: disable=no-member
# pylint: disable=too-many-locals
"""
Update source / receiving / high volume site CSV and Layer in AGOL
"""
import json
import os
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError
from copy import deepcopy
import logging
from arcgis.gis import GIS
from arcgis.features import FeatureLayerCollection, FeatureLayer
import openpyxl
import constant
import helper

MAPHUB_USER = os.getenv('MAPHUB_USER')
MAPHUB_PASS = os.getenv('MAPHUB_PASS')
SRC_CSV_ID = os.getenv('SRC_CSV_ID')
SRC_LAYER_ID = os.getenv('SRC_LAYER_ID')
RCV_CSV_ID = os.getenv('RCV_CSV_ID')
RCV_LAYER_ID = os.getenv('RCV_LAYER_ID')
HV_CSV_ID = os.getenv('HV_CSV_ID')
HV_LAYER_ID = os.getenv('HV_LAYER_ID')
LOGLEVEL = os.getenv('LOGLEVEL')
TIMEOUT_SECONDS = os.getenv('TIMEOUT_SECONDS')

# If there is no TIMEOUT_SECONDS, set it to 300 seconds (5 minutes)
if not TIMEOUT_SECONDS:
    TIMEOUT_SECONDS = '300'

config = helper.read_config()
MAPHUB_URL = config['AGOL']['MAPHUB_URL']

logging.basicConfig(level=LOGLEVEL, format='%(asctime)s [%(levelname)s] %(message)s')

def run_with_timeout(func, *args, timeout=TIMEOUT_SECONDS, **kwargs):
    """Run a function with a timeout and log an error if it exceeds the timeout."""
    # Ensure timeout is a number (convert from string if needed)
    if isinstance(timeout, str):
        try:
            timeout = int(timeout)  # Convert to integer if it's a string
        except ValueError:
            logging.error("Invalid timeout value: %s. Must be a number.", timeout)
            return None

    with ThreadPoolExecutor() as executor:
        future = executor.submit(func, *args, **kwargs)
        try:
            # Wait for the function to complete with the specified timeout
            return future.result(timeout=timeout)
        except FuturesTimeoutError:
            # Log an error with the specific function name if the operation times out
            logging.error("Operation '%s' timed out after %s seconds.", func.__name__, timeout)
            os._exit(1)

def connect_to_agol():
    """Connect to AGOL GIS with timeout handling."""
    logging.info('Connecting to AGOL GIS...')
    return run_with_timeout(GIS, MAPHUB_URL, username=MAPHUB_USER, password=MAPHUB_PASS)

def agol_items_overwrite():
    """Update source / receiving / high volume site CSV and Layer in AGOL"""
    # Run connection with timeout
    _gis = connect_to_agol()

    # If connection fails, skip the update processes
    if _gis is None:
        logging.error("Failed to connect to AGOL. Skipping update processes.")
        return

    # Proceed with updates, even if one of them times out
    # Timeouts are handled inside the update functions
    logging.info("Starting update of Source Site...")
    run_with_timeout(update_source_site, _gis)

    logging.info("Starting update of Receiving Site...")
    run_with_timeout(update_receiving_site, _gis)

    logging.info("Starting update of High Volume Site...")
    run_with_timeout(update_high_volume_site, _gis)

def update_source_site(_gis):
    """Update Source Site CSV and Feature Layer"""
    logging.info('Updating Soil Relocation Soruce Site CSV...')
    _src_csv_item = run_with_timeout(_gis.content.get, SRC_CSV_ID)
    if _src_csv_item is None:
        logging.error('Source Site CSV Item ID is invalid!')
        return

    _src_csv_update_result = run_with_timeout(_src_csv_item.update, {}, constant.SOURCE_CSV_FILE)
    if _src_csv_update_result is None:
        logging.error('Timeout occurred while updating Source Site CSV.')
        return

    logging.info("Updated Soil Relocation Source Site CSV successfully:%s", str(_src_csv_update_result))

    logging.info('Updating Soil Relocation Soruce Site Feature Layer...')

    # Wrap the overwrite operation with timeout
    _src_lyr_item = run_with_timeout(_gis.content.get, SRC_LAYER_ID)
    if _src_lyr_item is None:
        logging.error('Source Site Layer Item ID is invalid!')
        return

    _src_flc = FeatureLayerCollection.fromitem(_src_lyr_item)
    _src_lyr_overwrite_result = run_with_timeout(_src_flc.manager.overwrite, constant.SOURCE_CSV_FILE)

    if _src_lyr_overwrite_result is None:
        logging.error("Timeout occurred while updating Source Site Feature Layer.")
        return  # Stop further execution if timeout happens

    logging.info("Updated Soil Relocation Source Site Feature Layer successfully:%s", json.dumps(_src_lyr_overwrite_result))

    logging.info('Adding Soil Relocation Soruce Site Feature Layer Fields Description...')

    # Wrap the field description update with timeout
    fields_desc_list = get_fields_desc_list(constant.SOURCE_FIELDS_DESC_FILE)
    search = run_with_timeout(_gis.content.search, "id:" + SRC_LAYER_ID, timeout=TIMEOUT_SECONDS, item_type="Feature Layer")

    if search is None:
        logging.error("Timeout occurred while searching for Source Site Feature Layer.")
        return  # Stop further execution if timeout happens

    run_with_timeout(add_layer_fields_desc, search, fields_desc_list)

def update_receiving_site(_gis):
    """Update Receiving Site CSV and Feature Layer"""
    logging.info('Updating Soil Relocation Receiving Site CSV...')
    _rcv_csv_item = run_with_timeout(_gis.content.get, RCV_CSV_ID)
    if _rcv_csv_item is None:
        logging.error('Receiving Site CSV Item ID is invalid!')
        return

    _rcv_csv_update_result = run_with_timeout(_rcv_csv_item.update, {}, constant.RECEIVE_CSV_FILE)
    if _rcv_csv_update_result is None:
        logging.error('Timeout occurred while updating Receiving Site CSV.')
        return

    logging.info("Updated Soil Relocation Receiving Site CSV successfully:%s", str(_rcv_csv_update_result))

    logging.info('Updating Soil Relocation Receiving Site Feature Layer...')
    _rcv_lyr_item = run_with_timeout(_gis.content.get, RCV_LAYER_ID)
    if _rcv_lyr_item is None:
        logging.error('Receiving Site Layer Item ID is invalid!')
        return

    _rcv_flc = FeatureLayerCollection.fromitem(_rcv_lyr_item)
    _rcv_lyr_overwrite_result = run_with_timeout(_rcv_flc.manager.overwrite, constant.RECEIVE_CSV_FILE)

    if _rcv_lyr_overwrite_result is None:
        logging.error("Timeout occurred while updating Receiving Site Feature Layer.")
        return  # Stop further execution if timeout happens

    logging.info("Updated Soil Relocation Receiving Site Feature Layer successfully:%s", json.dumps(_rcv_lyr_overwrite_result))

    logging.info('Adding Soil Relocation Receiving Site Feature Layer Fields Description...')
    fields_desc_list = get_fields_desc_list(constant.RECEIVE_FIELDS_DESC_FILE)
    search = run_with_timeout(_gis.content.search, "id:" + RCV_LAYER_ID, timeout=TIMEOUT_SECONDS, item_type="Feature Layer")

    if search is None:
        logging.error("Timeout occurred while searching for Receiving Site Feature Layer.")
        return  # Stop further execution if timeout happens

    run_with_timeout(add_layer_fields_desc, search, fields_desc_list)

def update_high_volume_site(_gis):
    """Update High Volume Site CSV and Feature Layer"""
    logging.info('Updating High Volume Receiving Site CSV...')
    _hv_csv_item = run_with_timeout(_gis.content.get, HV_CSV_ID)
    if _hv_csv_item is None:
        logging.error('High Volume Receiving Site CSV Item ID is invalid!')
        return

    _hv_csv_update_result = run_with_timeout(_hv_csv_item.update, {}, constant.HIGH_VOLUME_CSV_FILE)
    if _hv_csv_update_result is None:
        logging.error('Timeout occurred while updating High Volume Site CSV.')
        return

    logging.info("Updated High Volume Receiving Site CSV successfully: %s", str(_hv_csv_update_result))

    logging.info('Updating High Volume Receiving Site Feature Layer...')
    _hv_lyr_item = run_with_timeout(_gis.content.get, HV_LAYER_ID)
    if _hv_lyr_item is None:
        logging.error('High Volume Receiving Site Layer Item ID is invalid!')
        return

    _hv_flc = FeatureLayerCollection.fromitem(_hv_lyr_item)
    _hv_lyr_overwrite_result = run_with_timeout(_hv_flc.manager.overwrite, constant.HIGH_VOLUME_CSV_FILE)

    if _hv_lyr_overwrite_result is None:
        logging.error("Timeout occurred while updating High Volume Site Feature Layer.")
        return  # Stop further execution if timeout happens

    logging.info("Updated High Volume Receiving Site Feature Layer successfully:%s", json.dumps(_hv_lyr_overwrite_result))

    logging.info('Adding High Volume Receiving Site Feature Layer Fields Description...')
    fields_desc_list = get_fields_desc_list(constant.HIGH_VOLUME_FIELDS_CSV_FILE)
    search = run_with_timeout(_gis.content.search, "id:" + HV_LAYER_ID, timeout=TIMEOUT_SECONDS, item_type="Feature Layer")

    if search is None:
        logging.error("Timeout occurred while searching for High Volume Site Feature Layer.")
        return  # Stop further execution if timeout happens

    run_with_timeout(add_layer_fields_desc, search, fields_desc_list)

def add_layer_fields_desc(search, fields_desc_list):
    """Add feature layer fields description in AGOL"""

    feature_layer = FeatureLayer.fromitem(search[0], layer_id=0)
    layer_fields = feature_layer.manager.properties.fields

    update_json = []
    for field in layer_fields:
        field_name = field['name']
        for lookup_field in fields_desc_list:
            # see if they match a field in the excel document
            if lookup_field[0] == field_name:
                # store the field JSON from the online layer
                field_json = dict(deepcopy(field))
                # assign the new field description in JSON format, if specified
                if lookup_field[1]:
                    long_desc = lookup_field[1]
                    # remove escape characters like double quotes, newlines, or encoding issues
                    if "<" in long_desc or ">" in long_desc:
                        logging.debug("Special character '>' or '<' found in field: %s", field_name)
                        logging.debug("Script will not run as expected. Please remove all hyperlinks or > < characters from the long description and rerun the script.")
                    long_desc = long_desc.replace('"', '\\\"').replace("\n", " ").replace("\t", " ").replace('\xa0', ' ').replace(">=", " greater than or equal to ").replace("<=", " less than or equal to ").replace(">", " greater than ").replace("<", " less than ")
                else:
                    long_desc = ""
                # build the JSON structure with the proper backslashes and quotes
                field_json['description'] = "{\"value\":" + "\"" + long_desc + "\"" + ",\"fieldValueType\":\"\"}"
                # field_json.pop('sqlType')
                if long_desc != "":
                    logging.debug("The long description for this field was also updated")
                # create a python list containing fields to update
                update_json.append(field_json)

    if update_json:
        logging.debug("Updating layer fields description in AGOL ...")
        #json_format = json.dumps(update_json)
        alias_update_dict = {'fields': update_json}
        #alias_update_json = json.dumps(alias_update_dict)
        response = feature_layer.manager.update_definition(alias_update_dict)
        logging.info("Feature Layer Fields Description Addition Task Complete:%s", response)

def get_fields_desc_list(fields_desc_file_name):
    """Get fields and descriptions from excel document"""
    xlsx = os.path.normpath(r"./"+constant.FIELDS_DESC_PATH+"/"+fields_desc_file_name)

    workbook = openpyxl.load_workbook(xlsx)
    sheet = workbook.active

    fields_desc_list = []

    iter_rows = sheet.iter_rows()
    next(iter_rows)
    for row in iter_rows:
        inner_list = []
        for val in row:
            inner_list.append(val.value)
        fields_desc_list.append(inner_list)

    return fields_desc_list
